"""Seed the AI Economy Universe database from the source xlsx file.

Usage:
    python seed.py --xlsx /path/to/AI_Economy_Universe_Final.xlsx

The script runs in two passes:
  Pass 1 - populate universe_taxonomy (categories then subcategories).
  Pass 2 - group rows by company_name, call Claude for multi-category
            reason where needed, then insert universe_companies.

Any row that cannot be inserted is written to seed_errors.log and skipped.
The script prints a summary on exit.
"""
import argparse
import logging
import os
import sys
from collections import defaultdict

import openpyxl
from openai import OpenAI
from dotenv import load_dotenv

load_dotenv()

# Make shared db_utils importable when run directly
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "../../shared/src"))

import db  # noqa: E402  (must come after sys.path patch)
from db import get_db, init_db  # noqa: E402

SHEET_NAME = "AI Economy Universe"
COLUMNS = ("category", "subcategory", "company_name", "ticker", "market", "country", "website")

logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)s %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("seed_errors.log", mode="w"),
    ],
)
log = logging.getLogger(__name__)


# --------------------------------------------------------------------------- #
# Helpers                                                                      #
# --------------------------------------------------------------------------- #

def _read_xlsx(path: str) -> list[dict]:
    """Return all non-empty data rows from the universe sheet."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found in {path}")
    ws = wb[SHEET_NAME]
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = dict(zip(COLUMNS, row))
        if not any(values.values()):
            continue
        # Validate required fields
        missing = [k for k in COLUMNS if not values.get(k)]
        if missing:
            log.error("Row %d: missing fields %s - skipped", i, missing)
            continue
        values["_row"] = i
        rows.append(values)
    wb.close()
    return rows


def _seed_taxonomy(rows: list[dict], conn) -> tuple[dict[str, int], dict[str, int]]:
    """Insert all categories and subcategories. Return name→id maps."""
    # Collect ordered unique values preserving first-seen order
    seen_cats: dict[str, None] = {}
    seen_subs: dict[tuple[str, str], None] = {}
    for r in rows:
        seen_cats[r["category"]] = None
        seen_subs[(r["category"], r["subcategory"])] = None

    cat_id: dict[str, int] = {}
    sub_id: dict[str, int] = {}

    # Insert categories
    for name in seen_cats:
        try:
            cur = conn.execute(
                """
                INSERT INTO universe_taxonomy (type, name)
                VALUES ('category', :name)
                ON CONFLICT (type, name) DO UPDATE SET name = EXCLUDED.name
                RETURNING id
                """,
                {"name": name},
            )
            cat_id[name] = cur.fetchone()["id"]
        except Exception as exc:
            log.error("taxonomy category '%s': %s", name, exc)

    # Insert subcategories linked to their parent category
    for (cat_name, sub_name) in seen_subs:
        parent = cat_id.get(cat_name)
        if parent is None:
            log.error("subcategory '%s': parent category '%s' not found - skipped", sub_name, cat_name)
            continue
        try:
            cur = conn.execute(
                """
                INSERT INTO universe_taxonomy (type, name, parent_id)
                VALUES ('subcategory', :name, :parent_id)
                ON CONFLICT (type, name) DO UPDATE SET parent_id = EXCLUDED.parent_id
                RETURNING id
                """,
                {"name": sub_name, "parent_id": parent},
            )
            sub_id[sub_name] = cur.fetchone()["id"]
        except Exception as exc:
            log.error("taxonomy subcategory '%s': %s", sub_name, exc)

    log.info("Taxonomy: %d categories, %d subcategories", len(cat_id), len(sub_id))
    return cat_id, sub_id


def _llm_client() -> OpenAI:
    return OpenAI(
        api_key=os.environ["OPENROUTER_API_KEY"],
        base_url="https://openrouter.ai/api/v1",
    )


def _generate_multi_category_reason(company_name: str, categories: list[str]) -> str:
    """Ask Claude via OpenRouter why this company spans multiple categories."""
    cat_list = "\n".join(f"  • {c}" for c in categories)
    response = _llm_client().chat.completions.create(
        model=os.environ.get("OPENROUTER_MODEL", "anthropic/claude-sonnet-4-6"),
        max_tokens=200,
        temperature=0.2,
        messages=[{
            "role": "user",
            "content": (
                f"You are an AI economy analyst maintaining a curated universe of companies.\n\n"
                f"The company **{company_name}** has been classified under multiple categories:\n"
                f"{cat_list}\n\n"
                f"Write a concise 1–2 sentence explanation of why {company_name} legitimately "
                f"spans these categories. Be specific about the distinct business lines or "
                f"products that justify each category. Return only the explanation, no preamble."
            ),
        }],
    )
    return (response.choices[0].message.content or "").strip()


def _seed_companies(
    rows: list[dict],
    cat_id: dict[str, int],
    sub_id: dict[str, int],
    conn,
) -> tuple[int, int]:
    """Group rows by company_name and insert one record per company."""
    # Group rows by normalised company name
    by_company: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        by_company[r["company_name"]].append(r)

    inserted = 0
    skipped = 0

    for company_name, company_rows in by_company.items():
        # Resolve category and subcategory IDs for each row
        resolved_cat_ids = []
        resolved_sub_ids = []
        valid = True

        for r in company_rows:
            cid = cat_id.get(r["category"])
            sid = sub_id.get(r["subcategory"])
            if cid is None or sid is None:
                log.error(
                    "Row %d (%s): unresolved taxonomy (cat=%s, sub=%s) - skipped",
                    r["_row"], company_name, r["category"], r["subcategory"],
                )
                valid = False
                break
            if cid not in resolved_cat_ids:
                resolved_cat_ids.append(cid)
            if sid not in resolved_sub_ids:
                resolved_sub_ids.append(sid)

        if not valid:
            skipped += 1
            continue

        # Use the first row for scalar fields (ticker, market, country, website
        # are consistent across rows for the same company)
        r0 = company_rows[0]

        multi_reason = None
        if len(resolved_cat_ids) > 1:
            categories = [r["category"] for r in company_rows]
            log.info("Generating multi-category reason for '%s'...", company_name)
            try:
                multi_reason = _generate_multi_category_reason(company_name, categories)
                log.info("  → %s", multi_reason)
            except Exception as exc:
                log.error("Claude call failed for '%s': %s - using placeholder", company_name, exc)
                multi_reason = "Multi-category: reason pending review."

        try:
            conn.execute(
                """
                INSERT INTO universe_companies (
                    company_name, ticker, market, country, website,
                    category_ids, subcategory_ids, multi_category_reason,
                    status, agent_added
                ) VALUES (
                    :company_name, :ticker, :market, :country, :website,
                    :category_ids, :subcategory_ids, :multi_category_reason,
                    'verified', FALSE
                )
                ON CONFLICT (company_name) DO NOTHING
                """,
                {
                    "company_name": company_name,
                    "ticker": r0["ticker"],
                    "market": r0["market"],
                    "country": r0["country"],
                    "website": r0["website"],
                    "category_ids": resolved_cat_ids,
                    "subcategory_ids": resolved_sub_ids,
                    "multi_category_reason": multi_reason,
                },
            )
            inserted += 1
        except Exception as exc:
            log.error("INSERT '%s': %s - skipped", company_name, exc)
            skipped += 1

    return inserted, skipped


# --------------------------------------------------------------------------- #
# Entrypoint                                                                   #
# --------------------------------------------------------------------------- #

def main() -> None:
    parser = argparse.ArgumentParser(description="Seed the AI Economy Universe database.")
    parser.add_argument(
        "--xlsx",
        default=os.environ.get("UNIVERSE_XLSX_PATH", "AI_Economy_Universe_Final.xlsx"),
        help="Path to the source xlsx file.",
    )
    args = parser.parse_args()

    log.info("=== research-universe seed ===")
    log.info("Source: %s", args.xlsx)

    log.info("Initialising database schema...")
    init_db()

    log.info("Reading xlsx...")
    rows = _read_xlsx(args.xlsx)
    log.info("Read %d data rows", len(rows))

    with get_db() as conn:
        log.info("Pass 1 - seeding taxonomy...")
        cat_id, sub_id = _seed_taxonomy(rows, conn)

        log.info("Pass 2 - seeding companies...")
        inserted, skipped = _seed_companies(rows, cat_id, sub_id, conn)

    log.info("")
    log.info("=== Seed complete ===")
    log.info("Companies inserted : %d", inserted)
    log.info("Rows skipped       : %d", skipped)
    if skipped:
        log.info("See seed_errors.log for details.")


if __name__ == "__main__":
    main()
