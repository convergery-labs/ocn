"""One-time bootstrap: read XLSX → create Google Sheet with all tabs."""
from __future__ import annotations

import datetime

import openpyxl

from sheets.client import _make_gc

EXISTING_TABS = [
    "Research Universe",
    "Public Companies",
    "Private Watchlist",
    "Category Guide",
    "Ticker QA",
]

DISCOVERY_QUEUE_HEADERS = [
    "Run ID",
    "Date Added",
    "Source",
    "Company Name",
    "Ticker",
    "Exchange",
    "Country / Market",
    "Source Description",
    "Seed URL",
    "Official URL",
    "Page Summary",
    "URL Status",
    "Summary Status",
    "Duplicate Status",
    "Possible Duplicate",
    "Review Note",
    "Row Status",
    "Category",
    "Subcategory",
    "Company Type",
    "Universe Type",
    "Classifier Reason",
]

NEW_COMPANIES_HEADERS = [
    "Category",
    "Subcategory",
    "Company",
    "Ticker / Symbol / Identifier",
    "Type",
    "Country / Market",
    "Source URLs",
    "Run ID",
    "Classified Date",
]

CHANGE_LOG_HEADERS = [
    "Date",
    "Run ID",
    "Action",
    "Company",
    "Category",
    "Type",
    "Reason",
    "Source URLs",
]


def _cell_str(cell) -> str:
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()
    return str(v)


def run_init_sheet(xlsx_path: str) -> str:
    gc = _make_gc()
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    spreadsheet = gc.create("AlphaStreet AI Research Universe")

    first = True
    for tab_name in EXISTING_TABS:
        if tab_name not in wb.sheetnames:
            continue
        rows = [[_cell_str(c) for c in row] for row in wb[tab_name].iter_rows()]
        if not rows:
            continue
        if first:
            ws = spreadsheet.sheet1
            ws.update_title(tab_name)
            first = False
        else:
            ws = spreadsheet.add_worksheet(
                title=tab_name, rows=max(len(rows) + 10, 100), cols=26
            )
        ws.update(values=rows, range_name="A1")

    wb.close()

    for title, headers in [
        ("Discovery Queue", DISCOVERY_QUEUE_HEADERS),
        ("New Companies", NEW_COMPANIES_HEADERS),
        ("Change Log", CHANGE_LOG_HEADERS),
    ]:
        ws = spreadsheet.add_worksheet(title=title, rows=1000, cols=len(headers))
        ws.update(values=[headers], range_name="A1")

    return spreadsheet.id
